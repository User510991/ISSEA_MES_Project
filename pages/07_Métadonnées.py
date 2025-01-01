import streamlit as st
import openpyxl
import requests
from openpyxl import load_workbook

# Titre de l'application Streamlit
st.title("Tableau des métadonnées")

# Charger le fichier Excel



# URL du fichier
url = "https://raw.githubusercontent.com/User510991/ISSEA_MES_Project/main/Livre.xlsx"

# Télécharger le fichier et l'enregistrer localement
response = requests.get(url)
if response.status_code == 200:
    with open("Livre.xlsx", "wb") as file:
        file.write(response.content)
    print("Fichier téléchargé avec succès.")
else:
    print(f"Échec du téléchargement. Code HTTP : {response.status_code}")
    exit()

# Charger le fichier avec openpyxl
try:
    workbook = load_workbook("Livre.xlsx")
    print("Workbook chargé avec succès.")
except Exception as e:
    print(f"Erreur lors du chargement du fichier : {e}")
        

    # Charger le fichier Excel avec openpyxl

    # Sélectionner la première feuille (ou une feuille spécifique)
sheet = workbook.active  # Ou sheet = workbook['Nom_de_la_feuille']

    # Créer une liste pour stocker les données du tableau
data = []

    # Lire les lignes de la feuille Excel
for row in sheet.iter_rows(values_only=True):
  data.append(row)

    # Afficher les données dans un tableau Streamlit
st.table(data)
